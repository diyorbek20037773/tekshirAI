"""Admin Excel upload API — maktab/sinf/o'qituvchi/o'quvchi ma'lumotlarini yuklash.

Jarayoni:
1. GET /api/admin/upload/template/{type} — bo'sh shablon yuklab olish
2. POST /api/admin/upload/{type}/preview — yuklangan Excel parse + validatsiya, preview qaytaradi
3. POST /api/admin/upload/{type}/confirm — DB ga saqlash (preview tasdiqlangach)
4. GET /api/admin/upload/stats — statistika (qancha maktab, o'qituvchi, o'quvchi)
"""

import logging
import os
from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import load_workbook
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import get_db
from backend.api.admin import verify_admin_token
from backend.models.directory import (
    School, ClassroomDirectory, TeacherDirectory, StudentDirectory,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/admin/upload", tags=["admin-upload"], dependencies=[Depends(verify_admin_token)])

# Shablon fayllar joylashuvi
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "shablon_excel")

TEMPLATE_FILES = {
    "maktablar": "1_maktablar.xlsx",
    "sinflar": "2_sinflar.xlsx",
    "oqituvchilar": "3_oqituvchilar.xlsx",
    "oquvchilar": "4_oquvchilar.xlsx",
}


# ============================================================
# 1) SHABLON YUKLAB OLISH
# ============================================================

@router.get("/template/{type}")
async def download_template(type: Literal["maktablar", "sinflar", "oqituvchilar", "oquvchilar"]):
    """Bo'sh Excel shablonni yuklab olish."""
    filename = TEMPLATE_FILES.get(type)
    if not filename:
        raise HTTPException(status_code=404, detail="Bunday shablon yo'q")
    path = os.path.join(TEMPLATE_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=500, detail=f"Shablon fayli topilmadi: {filename}")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


# ============================================================
# 2) EXCEL PARSE + VALIDATSIYA (preview)
# ============================================================

def _parse_excel(file_bytes: bytes) -> list[list]:
    """Excel faylni qator-qator ro'yxat qilib qaytarish (sarlavhasiz)."""
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # sarlavha
        # Bo'sh qatorlarni o'tkazib yuborish
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        rows.append(list(row))
    return rows


def _str(v):
    """Cell qiymatini matnga o'tkazish."""
    if v is None:
        return ""
    return str(v).strip()


def _int(v, default=None):
    try:
        return int(v) if v is not None else default
    except (ValueError, TypeError):
        return default


@router.post("/{type}/preview")
async def preview_upload(
    type: Literal["maktablar", "sinflar", "oqituvchilar", "oquvchilar"],
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Excelni parse qilib, validatsiya qiladi va preview qaytaradi (DB ga yozmaydi)."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Faqat .xlsx fayl qabul qilinadi")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fayl 5MB dan katta")

    try:
        rows = _parse_excel(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel o'qishda xato: {str(e)[:200]}")

    valid = []
    errors = []

    if type == "maktablar":
        for i, row in enumerate(rows, start=2):
            row += [None] * (4 - len(row))
            viloyat, tuman, name, address = row[:4]
            v, t, n = _str(viloyat), _str(tuman), _str(name)
            if not v or not t or not n:
                errors.append({"row": i, "error": "Viloyat, Tuman va Maktab nomi majburiy"})
                continue
            valid.append({"viloyat": v, "tuman": t, "name": n, "address": _str(address)})

    elif type == "sinflar":
        for i, row in enumerate(rows, start=2):
            row += [None] * (4 - len(row))
            school_name, cls_name, grade, subject = row[:4]
            sn, cn, sj = _str(school_name), _str(cls_name), _str(subject)
            g = _int(grade)
            if not sn or not cn or not g or not sj:
                errors.append({"row": i, "error": "Maktab, Sinf nomi, raqami va Fan majburiy"})
                continue
            if g < 1 or g > 11:
                errors.append({"row": i, "error": f"Sinf raqami 1-11 oralig'ida bo'lishi kerak (sizniki: {g})"})
                continue
            valid.append({"school_name": sn, "name": cn, "grade": g, "subject": sj})

    elif type == "oqituvchilar":
        for i, row in enumerate(rows, start=2):
            row += [None] * (7 - len(row))
            school_name, last, first, middle, subject, phone, sinf_rahbari = row[:7]
            sn, ln, fn, sj = _str(school_name), _str(last), _str(first), _str(subject)
            if not sn or not ln or not fn or not sj:
                errors.append({"row": i, "error": "Maktab, Familiya, Ism, Fan majburiy"})
                continue
            valid.append({
                "school_name": sn, "last_name": ln, "first_name": fn,
                "middle_name": _str(middle), "subject": sj,
                "phone": _str(phone), "sinf_rahbari": _str(sinf_rahbari),
            })

    elif type == "oquvchilar":
        for i, row in enumerate(rows, start=2):
            row += [None] * (7 - len(row))
            school_name, classroom, last, first, middle, byear, gender = row[:7]
            sn, cn, ln, fn = _str(school_name), _str(classroom), _str(last), _str(first)
            if not sn or not cn or not ln or not fn:
                errors.append({"row": i, "error": "Maktab, Sinf, Familiya, Ism majburiy"})
                continue
            # Sinfdan grade ajratish ("7-A" → 7)
            grade = None
            try:
                grade = int(cn.split("-")[0])
            except (ValueError, IndexError):
                errors.append({"row": i, "error": f"Sinf format noto'g'ri (kutilgan: '7-A', sizniki: '{cn}')"})
                continue
            valid.append({
                "school_name": sn, "classroom_name": cn, "grade": grade,
                "last_name": ln, "first_name": fn, "middle_name": _str(middle),
                "birth_year": _int(byear), "gender": _str(gender) or None,
            })

    return {
        "type": type,
        "total_rows": len(rows),
        "valid_count": len(valid),
        "error_count": len(errors),
        "errors": errors[:50],  # Eng ko'pi 50 ta xato ko'rsatamiz
        "preview": valid[:10],  # Birinchi 10 ta yozuv preview
        "valid_data": valid,  # Confirm uchun (frontend body'da qaytaradi)
    }


# ============================================================
# 3) DB GA SAQLASH (confirm)
# ============================================================

from pydantic import BaseModel
from typing import Any


class ConfirmRequest(BaseModel):
    valid_data: list[dict[str, Any]]
    replace: bool = False  # True bo'lsa, eski ma'lumotlar o'chiriladi


@router.post("/{type}/confirm")
async def confirm_upload(
    type: Literal["maktablar", "sinflar", "oqituvchilar", "oquvchilar"],
    body: ConfirmRequest,
    db: AsyncSession = Depends(get_db),
):
    """Validatsiyadan o'tgan ma'lumotlarni DB ga yozish."""
    inserted = 0
    skipped = 0
    errors = []

    try:
        if type == "maktablar":
            if body.replace:
                await db.execute(delete(School))
            for item in body.valid_data:
                # Mavjud bo'lsa o'tkazib yuboramiz (UNIQUE constraint)
                exists = await db.execute(select(School).where(School.name == item["name"]))
                if exists.scalars().first():
                    skipped += 1
                    continue
                db.add(School(
                    name=item["name"], viloyat=item["viloyat"],
                    tuman=item["tuman"], address=item.get("address") or None,
                ))
                inserted += 1

        elif type == "sinflar":
            if body.replace:
                await db.execute(delete(ClassroomDirectory))
            # Maktablarni map qilish
            schools_res = await db.execute(select(School))
            schools_by_name = {s.name: s for s in schools_res.scalars().all()}
            for item in body.valid_data:
                school = schools_by_name.get(item["school_name"])
                if not school:
                    errors.append(f"Maktab topilmadi: {item['school_name']}")
                    continue
                exists = await db.execute(
                    select(ClassroomDirectory).where(
                        ClassroomDirectory.school_id == school.id,
                        ClassroomDirectory.name == item["name"],
                        ClassroomDirectory.subject == item["subject"],
                    )
                )
                if exists.scalars().first():
                    skipped += 1
                    continue
                db.add(ClassroomDirectory(
                    school_id=school.id, name=item["name"],
                    grade=item["grade"], subject=item["subject"],
                ))
                inserted += 1

        elif type == "oqituvchilar":
            if body.replace:
                await db.execute(delete(TeacherDirectory))
            schools_res = await db.execute(select(School))
            schools_by_name = {s.name: s for s in schools_res.scalars().all()}
            for item in body.valid_data:
                school = schools_by_name.get(item["school_name"])
                if not school:
                    errors.append(f"Maktab topilmadi: {item['school_name']}")
                    continue
                db.add(TeacherDirectory(
                    school_id=school.id,
                    last_name=item["last_name"], first_name=item["first_name"],
                    middle_name=item.get("middle_name") or None,
                    subject=item["subject"], phone=item.get("phone") or None,
                    sinf_rahbari=item.get("sinf_rahbari") or None,
                ))
                inserted += 1

        elif type == "oquvchilar":
            if body.replace:
                await db.execute(delete(StudentDirectory))
            schools_res = await db.execute(select(School))
            schools_by_name = {s.name: s for s in schools_res.scalars().all()}
            for item in body.valid_data:
                school = schools_by_name.get(item["school_name"])
                if not school:
                    errors.append(f"Maktab topilmadi: {item['school_name']}")
                    continue
                db.add(StudentDirectory(
                    school_id=school.id,
                    classroom_name=item["classroom_name"], grade=item["grade"],
                    last_name=item["last_name"], first_name=item["first_name"],
                    middle_name=item.get("middle_name") or None,
                    birth_year=item.get("birth_year"), gender=item.get("gender"),
                ))
                inserted += 1

        await db.flush()

    except Exception as e:
        await db.rollback()
        logger.error(f"Confirm upload xato ({type}): {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"DB xato: {str(e)[:200]}")

    return {
        "status": "ok",
        "inserted": inserted,
        "skipped": skipped,
        "errors": errors[:20],
    }


# ============================================================
# 4) STATISTIKA
# ============================================================

@router.get("/stats")
async def upload_stats(db: AsyncSession = Depends(get_db)):
    """Direktoriya ma'lumotlari statistikasi."""
    schools_count = (await db.execute(select(func.count()).select_from(School))).scalar() or 0
    classes_count = (await db.execute(select(func.count()).select_from(ClassroomDirectory))).scalar() or 0
    teachers_count = (await db.execute(select(func.count()).select_from(TeacherDirectory))).scalar() or 0
    students_count = (await db.execute(select(func.count()).select_from(StudentDirectory))).scalar() or 0
    return {
        "schools": schools_count,
        "classes": classes_count,
        "teachers": teachers_count,
        "students": students_count,
    }


@router.delete("/{type}")
async def delete_all(
    type: Literal["maktablar", "sinflar", "oqituvchilar", "oquvchilar"],
    db: AsyncSession = Depends(get_db),
):
    """Tegishli direktoriyani butunlay o'chirish (CASCADE bilan)."""
    model_map = {
        "maktablar": School,
        "sinflar": ClassroomDirectory,
        "oqituvchilar": TeacherDirectory,
        "oquvchilar": StudentDirectory,
    }
    model = model_map[type]
    result = await db.execute(delete(model))
    await db.flush()
    return {"status": "ok", "deleted": result.rowcount or 0}
